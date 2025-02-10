from concurrent.futures import ThreadPoolExecutor
import logging
import time
import openai
import os
import numpy as np
import gspread  # Библиотека для работы с Google Sheets
from google.oauth2.service_account import Credentials
from telegram import (
    InlineKeyboardMarkup,
    Update,
    ReplyKeyboardMarkup,
    BotCommand,
    ReplyKeyboardRemove,
)
from telegram._inline.inlinekeyboardbutton import InlineKeyboardButton
from telegram.ext import (
    ApplicationBuilder,
    CommandHandler,
    ContextTypes,
    MessageHandler,
    filters,
)
from dotenv import load_dotenv
from pymilvus import connections, Collection, utility
from telegram.ext._handlers.callbackqueryhandler import CallbackQueryHandler
import tiktoken
import boto3  # Библиотека для работы с MinIO (S3 совместимое API)
from botocore.exceptions import NoCredentialsError
import re
import asyncio
from datetime import datetime
from google.oauth2.service_account import Credentials
import requests
from openpyxl import load_workbook  # работа с xlsx
from io import StringIO
from io import BytesIO

# Загрузка переменных окружения из файла .env
load_dotenv("keys_google_sheet.env")
load_dotenv("keys_gpt_telegram.env")
load_dotenv("keys_milvus.env")
load_dotenv("keys_minio.env")

TELEGRAM_BOT_TOKEN = os.getenv("TELEGRAM_BOT_TOKEN")  # Токен ТГ Бота
OPENAI_API_KEY = os.getenv("OPENAI_API_KEY")  # API токен OpenAI
MODEL_GPT_INT = os.getenv("MODEL_GPT_INT")  # Модель ИИ, с которой ведется диалог

MINIO_ACCESS_KEY = os.getenv("MINIO_ACCESS_KEY")  # Логин для подключенияMiniO
MINIO_SECRET_KEY = os.getenv("MINIO_SECRET_KEY")  # Пароль для подключения MiniO
MINIO_ENDPOINT = os.getenv("MINIO_ENDPOINT")  # IP и порт MiniO
MINIO_REGION_NAME = os.getenv("MINIO_REGION_NAME")  # Регион MiniO
MINIO_BUCKET_NAME = os.getenv("MINIO_BUCKET_NAME")  # Название Бакета MiniO
MINIO_FOLDER_DOCS_NAME_SPRAVOCHNIK = os.getenv(
    "MINIO_FOLDER_DOCS_NAME_SPRAVOCHNIK"
)  # Название Папки хранения Таблиц/Изображений Справочника инженеров
MINIO_FOLDER_DOCS_NAME_MANUAL = os.getenv(
    "MINIO_FOLDER_DOCS_NAME_MANUAL"
)  # Название Папки хранения Таблиц/Изображений Мануала
MINIO_FOLDER_LOGS_NAME = os.getenv(
    "MINIO_FOLDER_LOGS_NAME"
)  # Место, куда сохраняются логи контекста


MILVUS_DB_NAME_FIRST = os.getenv(
    "MILVUS_DB_NAME_FIRST"
)  # БД коллекций Милвуса c справочником
MILVUS_DB_NAME_SECOND = os.getenv(
    "MILVUS_DB_NAME_SECOND"
)  # БД коллекций Милвуса с мануалами
MILVUS_COLLECTION = os.getenv("MILVUS_COLLECTION")  # Коллекция Милвуса
MILVUS_HOST = os.getenv("MILVUS_HOST")  # IP Милвуса
MILVUS_PORT = os.getenv("MILVUS_PORT")  # Порт Милвуса
MILVUS_USER = os.getenv("MILVUS_USER")  # Логин Милвуса(БД)
MILVUS_PASSWORD = os.getenv("MILVUS_PASSWORD")  # Пароль Милвуса(БД)

SPREADSHEET_ID = os.getenv("SPREADSHEET_ID")  # ID Google Таблицы MODEL_GPT_INT

private_key = os.getenv("GOOGLE_PRIVATE_KEY")
if not private_key:
    raise ValueError("GOOGLE_PRIVATE_KEY is not set")
private_key = private_key.replace("\\n", "\n")

google_credentials = {  # Тут все ключи для работы API от гугл щит
    "type": os.getenv("GOOGLE_TYPE"),
    "project_id": os.getenv("GOOGLE_PROJECT_ID"),
    "private_key_id": os.getenv("GOOGLE_PRIVATE_KEY_ID"),
    "private_key": private_key,  # Экранирование переносов строк
    "client_email": os.getenv("GOOGLE_CLIENT_EMAIL"),
    "client_id": os.getenv("GOOGLE_CLIENT_ID"),
    "auth_uri": os.getenv("GOOGLE_AUTH_URI"),
    "token_uri": os.getenv("GOOGLE_TOKEN_URI"),
    "auth_provider_x509_cert_url": os.getenv("GOOGLE_AUTH_PROVIDER_CERT_URL"),
    "client_x509_cert_url": os.getenv("GOOGLE_CLIENT_CERT_URL"),
}

URL = f"https://api.telegram.org/bot{TELEGRAM_BOT_TOKEN}/getUpdates"
firts_message_from_tg_bot = "Привет!🖐 Я асистент для инженеров, перед тем как задать мне вопрос, выбери режим работы через команду '/metod'"

minio_folder_docs_name = MINIO_FOLDER_DOCS_NAME_SPRAVOCHNIK
milvus_collection_name = MILVUS_COLLECTION
# milvus_collection_name = MILVUS_COLLECTION_SPRAVOCHNIK

# Создаем пул потоков (до 10 одновременно)
executor = ThreadPoolExecutor(
    max_workers=10
)  #  <========================== Количество потоков

# Устанавливаем ключ OpenAI API
openai.api_key = OPENAI_API_KEY

# Настройка логирования
logging.basicConfig(
    format="%(asctime)s - %(name)s - %(levelname)s - %(message)s", level=logging.INFO
)

logger = logging.getLogger(__name__)

# Настройка Google Sheets API
credentials = Credentials.from_service_account_info(
    google_credentials, scopes=["https://www.googleapis.com/auth/spreadsheets"]
)
client = gspread.authorize(credentials)
sheet = client.open_by_key(SPREADSHEET_ID).sheet1
logger.info("Подключение к MiniO начато")
# Настройка MinIO клиента
s3_client = boto3.client(
    "s3",
    endpoint_url=MINIO_ENDPOINT,
    aws_access_key_id=MINIO_ACCESS_KEY,
    aws_secret_access_key=MINIO_SECRET_KEY,
    region_name=MINIO_REGION_NAME,
)
logger.info("Подключение к MiniO завершено")
# print(f'Логин "{MINIO_ACCESS_KEY}" для БД MiniO')  # Проверка LOG
# print(f'Пароль "{MINIO_SECRET_KEY}" для БД MiniO')  # Проверка PSWD


logger.info("Подключение к коллекциям Milvus начато")
# Подключаемся к Milvus с справочником
connections.connect(
    alias="default",
    host=MILVUS_HOST,
    port=MILVUS_PORT,
    db_name=MILVUS_DB_NAME_FIRST,
    user=MILVUS_USER,
    password=MILVUS_PASSWORD,
)

# Получаем список всех коллекций в базе данных
all_collections = utility.list_collections()  # <======== Загрузка всех коллекций


# Собираем эмбеддинги из всех активных коллекций
all_texts = []
all_embeddings = []
all_table_references = []
all_related_tables = []  # Новый список для related_table

count_e = 0
target_collection_name = "engrs_spravochnik"

try:
    # Проверяем, существует ли коллекция с указанным именем
    if not utility.has_collection(target_collection_name):
        raise ValueError(f"Коллекция '{target_collection_name}' не существует.")

    # Загружаем коллекцию
    collection = Collection(name=target_collection_name)
    if collection.num_entities > 0:
        # Загружаем данные из коллекции
        entities = collection.query(
            expr="id > 0",
            output_fields=["embedding", "text", "reference", "related_table"],
        )
        all_texts = [entity["text"] for entity in entities]
        all_embeddings = [entity["embedding"] for entity in entities]
        all_table_references = [entity["reference"] for entity in entities]
        all_related_tables = [entity.get("related_table", "") for entity in entities]

        logger.info(f"Коллекция '{target_collection_name}' успешно загруженаааа.")
        logger.info("Подключение к коллекциям Milvus завершено")

    else:
        logger.info(f"Коллекция '{target_collection_name}' пуста.")
except Exception as e:
    logger.info(f"Ошибка при загрузке коллекции '{target_collection_name}': {e}")

logger.info(f"-----------------------------------")
logger.info(f"| Коллекция справочника загружена |")
logger.info(f"-----------------------------------")


def check_openai_access(retry_delay=5):
    """
    Проверяет доступ к OpenAI, отправляя тестовый запрос.
    Если доступа нет, пробует несколько раз.
    """

    try:
        # Отправляем тестовый запрос (Embedding)
        response = openai.embeddings.create(
            input=["Проверка доступа к OpenAI"], model="text-embedding-ada-002"
        )

        if response:
            logger.info("✅ Доступ к OpenAI подтверждён!")
            return True

    except openai.AuthenticationError:
        logger.info("❌ Ошибка: Неверный API-ключ OpenAI.")
        return False

    except openai.RateLimitError:
        logger.info("⏳ Доступ к OpenAI ограничен, попробуем позже...")
        time.sleep(retry_delay)

    except openai.OpenAIError as e:
        logger.info(f"⚠ Ошибка API OpenAI: {e}")
        time.sleep(retry_delay)

    except Exception as e:
        logger.info(f"🚨 Неизвестная ошибка при доступе к OpenAI: {e}")
        time.sleep(retry_delay)

    return False


# Запускаем проверку
check_openai_access()


# Метод для создания эмбеддинга запроса пользователя
def create_embedding_for_query(query, update: Update):
    try:
        response = openai.embeddings.create(
            input=[query],
            model="text-embedding-ada-002",
            timeout=10,  # таймаут на получение ответа
        )
        return response.data[0].embedding
    # except openai.error.Timeout as e:
    #    print(f"Ошибка: Таймаут запроса - {e}")
    except Exception as e:
        logger.error(f"Ошибка: {e}")
        update.message.reply_text(
            f"Произошла ошибка при создании вектора вопроса: {str(e)}"
        )
        return None


# Метод поиска наиболее релевантных эмбеддингов
def find_most_similar(query_embedding, top_n=15):
    query_embedding_np = np.array([query_embedding], dtype=np.float32)
    similarities = np.dot(all_embeddings, query_embedding_np.T)
    most_similar_indices = np.argsort(similarities, axis=0)[::-1].flatten()

    unique_related_tables = set()  # Храним уже добавленные related_table
    filtered_texts = []
    filtered_refs = []
    filtered_related_tables = []

    for i in most_similar_indices:
        related_table = all_related_tables[i]  # Получаем связанный related_table

        # Если related_table уже встречался — пропускаем
        if related_table in unique_related_tables:
            continue

        # Добавляем уникальный related_table в результат
        unique_related_tables.add(related_table)
        filtered_texts.append(all_texts[i])
        filtered_refs.append(all_table_references[i])
        filtered_related_tables.append(related_table)

        # Если уже набрали нужное количество top_n, выходим
        if len(filtered_texts) >= top_n:
            break

    return filtered_texts, filtered_refs, filtered_related_tables


def search_in_milvus(query_embedding, top_n=10):
    """
    Ищет вектор query_embedding в указанной коллекции Milvus (collection_name),
    возвращая top_n наиболее похожих результатов.

    Параметры:
    -----------
    query_embedding: list или np.array
        Вектор, который мы ищем.
    collection_name: str
        Название коллекции в Milvus, где искать.
    top_n: int
        Сколько результатов вернуть (по умолчанию 10).

    Возвращает:
    -----------
    results: list
        Список объектов Hit; каждый содержит поля:
        - hit.score: оценка сходства
        - hit.entity: словарь (по ключам из output_fields)
    """
    # print(f"query_embedding - {query_embedding}")
    # Убедимся, что у нас np.float32
    if not isinstance(query_embedding, np.ndarray):
        query_embedding = np.array(query_embedding, dtype=np.float32)
    else:
        query_embedding = query_embedding.astype(np.float32)

    # Инициализируем объект Collection
    collection = Collection(name="Manuals")

    # Параметры поиска
    search_params = {
        "metric_type": "L2",  # Замените на "IP", "COSINE" и т.п. при необходимости
        "params": {"nprobe": 10},
    }

    # Выполняем поиск
    results = collection.search(
        data=[query_embedding.tolist()],  # Передаём список векторов
        anns_field="embedding",  # Поле, в котором хранятся эмбеддинги
        param=search_params,
        limit=top_n,
        output_fields=[
            "text",
            "manual_id",
            "embedding",
        ],  # Добавьте поля, которые нужны
    )
    # Возвращаем список Hit-объектов для нашего запроса.
    # results[0] — это список из top_n найденных элементов
    hits = results[0]
    manuals_score = []  # совпадение векторов
    manual_ids = []  # Текстовая часть соответствующая вектору
    manual_texts = []  # Название оригинального документа
    manuals_embeddings = []  # Найденные схожие вектора

    for hit_score in hits:
        manual_score = 1 - hit_score.score
        # print(manual_score)
        if manual_score == 1.0:
            manual_score = 100.0
        else:
            manual_score = round((manual_score), 2) * 100
        manuals_score.append(manual_score)
        # print("score:", manual_score)

    for hit_id in hits:
        manual_id = hit_id.entity.get("manual_id")
        manual_ids.append(manual_id)
        # print("manual_id:", manual_id)

    for hit_text in hits:
        manual_text = hit_text.entity.get("text")
        manual_texts.append(manual_text)
        # print("text:", manual_text)

    for hit_embedding in hits:
        manuals_embedding = hit_embedding.entity.get("embedding")
        manuals_embeddings.append(manuals_embedding)
        # print("embedding:", manuals_embedding)

    for col, txt, score in zip(manual_ids, manual_texts, manuals_score):
        print("Совпадение:", score)
        print("Название мануала:", col)
        print("Текстовый блок:", txt)

    """for hit in hits:
        print("score:", hit.score)
        print("manual_id:", hit.entity.get("manual_id"))
        print("text:", hit.entity.get("text"))
        print("embedding:", hit.entity.get("embedding"))"""

    return manual_ids, manual_texts, manuals_score


def find_most_similar_with_collections(context, query_embedding, top_n=10):
    """
    Находит наиболее релевантные вектора и возвращает:
    - Тексты,
    - Коллекции, из которых взяты вектора,
    - Описание коллекций.
    """
    user_data = context.user_data
    query_embedding_np = np.array([query_embedding], dtype=np.float32)
    embeddings = np.array(user_data.get("all_embeddings", []))
    collections_per_embedding = user_data.get("all_collections_per_embedding", [])

    if len(embeddings) == 0:
        return [], set()  # Пустой список текстов и коллекций

    # Рассчитываем сходство
    similarities = np.dot(embeddings, query_embedding_np.T)
    most_similar_indices = np.argsort(similarities, axis=0)[::-1][:top_n]

    relevant_collections = set()

    # Логирование найденных элементов
    logger.info("Найденные релевантные элементы:")
    for idx in most_similar_indices.flatten():
        relevant_collections.add(collections_per_embedding[idx])
        logger.info(
            f"Текст: {user_data['all_texts'][idx]}, "
            f"Сходство: {similarities[idx][0]}"
            f"Коллекция: {collections_per_embedding[idx]}"
        )
    user_data["all_collections"] = list(relevant_collections)

    return list(relevant_collections)


# Чтение содержимого таблицы из MinIO (S3 хранилища)
def read_table_from_minio(table_reference):
    """Читает таблицу из MinIO и возвращает её содержимое в виде текста."""
    try:
        response = s3_client.get_object(Bucket=MINIO_BUCKET_NAME, Key=table_reference)
        buffer = BytesIO(response["Body"].read())  # Считываем файл в память
        workbook = load_workbook(buffer)  # Открываем файл как xlsx
        sheet = workbook.active  # Используем первый лист

        # Преобразуем содержимое таблицы в строковый формат
        table_content = ""
        for row in sheet.iter_rows(values_only=True):
            row_content = " | ".join(map(str, row))  # Преобразуем каждую строку
            table_content += row_content + "\n"

        return table_content.strip()
    except NoCredentialsError as e:
        logger.error(f"Ошибка аутентификации в MinIO: {e}")
        return None
    except Exception as e:
        logger.error(f"Не удалось прочитать таблицу из MinIO: {e}")
        return None


# Метод для обработки команды /start
async def start(update: Update, context):

    user_id = update.message.from_user.id
    last_selected_mode = load_user_mode_from_sheet(user_id)
    if last_selected_mode:
        context.user_data["selected_method"] = last_selected_mode

    await update.message.reply_text(
        firts_message_from_tg_bot, reply_markup=ReplyKeyboardRemove()
    )


# Метод для обработки команды /info
async def info(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """
    Обработка команды /info.
    Отправляет информацию о боте.
    """
    info_message = (
        "Я Ассистент для инженеров.\n"
        "Чему я уже научился🧐:\n"
        "1️⃣ Искать информации по Справочнику Инженеров.\n"
        "2️⃣ Отправлять таблицы или рисунки из Справочника Инжнеров.\n"
        "Если ты готов, то выбери режим работы через команду ➡️/metod⬅️!"
    )
    await update.message.reply_text(info_message, reply_markup=ReplyKeyboardRemove)


# Метод для обработки команды /metod
async def metod(update: Update, context: ContextTypes.DEFAULT_TYPE):
    keyboard = [
        [InlineKeyboardButton("Поиск по справочнику", callback_data="engs_bot")],
        # [InlineKeyboardButton("Поиск мануалов", callback_data="manuals_engrs")],
    ]
    reply_markup = InlineKeyboardMarkup(keyboard)

    await update.message.reply_text(
        "Выберите метод работы Бота:", reply_markup=reply_markup
    )


# Метод подсчитывает токены для конкретного отрывка текста
def count_tokens(text):
    encoding = tiktoken.encoding_for_model("text-embedding-ada-002")
    tokens = encoding.encode(text)
    return len(tokens)


# Метод для записи вопроса пользователя в Google Таблицу
def save_user_question_to_sheet(
    user_message, gpt_response, user_tag, log_filename, handle_message_method
):
    next_row = len(sheet.get_all_values()) + 1  # Следующий номер строки
    sheet.update(
        f"A{next_row}:G{next_row}",
        [
            [
                next_row - 1,
                user_message,
                gpt_response,
                "",
                user_tag,
                log_filename,
                handle_message_method,
            ]
        ],
    )  # Запись номера теста, вопроса, ответа GPT, оценки (пусто), и тега пользователя


user_image_context = {}


# Метод приоритизации поиска релевантных данных
def filter_and_prioritize_context(
    most_similar_texts, most_similar_refs, most_similar_related_tables
):
    texts_and_tables = []
    images = []
    additional_contexts = []  # Для хранения дополнительного контекста
    added_tables = set()  # Для отслеживания уже добавленных таблиц

    # Разделяем объекты на тексты/таблицы и изображения
    for i, ref in enumerate(most_similar_refs):
        related_table = most_similar_related_tables[i]

        # Обработка таблиц
        if ref.endswith(".xlsx"):
            if ref not in added_tables:
                table_content = read_table_from_minio(
                    f"{minio_folder_docs_name}/{ref}"
                )  # сюда ничего не пиши
                if table_content:
                    # texts_and_tables.append(
                    #    (
                    #        f"Таблица аааа ({ref}):\n{table_content}",
                    #        ref,
                    #    )
                    # )
                    added_tables.add(ref)  # Сюда тоже ничего не пиши
                else:
                    logger.warning(f"Не удалось прочитать таблицу: {ref}")
        # Обработка текста с родительской таблицей
        elif related_table:
            # Проверяем, есть ли связь с таблицей
            if related_table not in added_tables:
                table_content = read_table_from_minio(
                    f"{minio_folder_docs_name}/{related_table}"
                )
                if table_content:
                    table_name = next(
                        (
                            text
                            for text, reference in zip(all_texts, all_table_references)
                            if reference == related_table
                        ),
                        "Безымянная таблица",
                    )
                    texts_and_tables.append(
                        (
                            f"Текстовый блок\n ({table_name}):\n{table_content} \nКонец текстового блока",
                            related_table,
                        )
                    )
                    added_tables.add(related_table)

            # Ищем дополнительные объекты, связанные с таблицей
            found_objects = search_by_reference_in_milvus(related_table)
            if found_objects:
                for obj in found_objects:
                    additional_contexts.append(obj["text"])

        # Обработка изображений
        else:
            images.append((most_similar_texts[i], ref))

    # Ограничиваем количество текстов и таблиц до 10
    prioritized_texts_and_tables = texts_and_tables[:10]

    # Ограничиваем количество изображений до 10
    prioritized_images = images[:10]

    # Возвращаем два отдельных списка и дополнительный контекст
    return prioritized_texts_and_tables, prioritized_images, additional_contexts


def search_by_reference_in_milvus(reference_value):
    """Ищет объекты в Milvus, у которых reference совпадает с указанным значением."""
    collection = Collection(name=milvus_collection_name)
    try:
        # Выполняем запрос к Milvus
        result = collection.query(
            expr=f'reference == "{reference_value}"',
            output_fields=["text", "reference"],
        )
        return result if result else None
    except Exception as e:
        logger.error(
            f"Ошибка при поиске в Milvus для reference '{reference_value}': {e}"
        )
        return None


# Самый главный метод, обработки, получения, отправления сообщений
async def handle_message(update: Update, context):
    user_id = update.message.from_user.id
    last_selected_mode = load_user_mode_from_sheet(user_id)
    # print("last_selected_mode", last_selected_mode)
    if last_selected_mode:
        context.user_data["last_selected_mode"] = last_selected_mode
    # Динамический вызов нужного метода обработки

    handle_message_method = context.user_data.get(
        "last_selected_mode",  # Если метод сохранён, используем его
        handle_message,  # По умолчанию - текущий метод
    )
    # print(f"handle_message_method - {handle_message_method}")

    # Если метод другой (например, handle_message_manuals), вызываем его
    if handle_message_method != "engs_bot":
        await handle_message_manuals(update, context)
        return

    user_id = update.message.from_user.id

    # Проверяем, ждет ли бот оценку
    if context.user_data.get("awaiting_feedback", False):
        await update.message.reply_text(
            "⚠️ Сначала оцените предыдущий ответ, прежде чем задать новый вопрос!"
        )
        return  # Блокируем новый вопрос

    user_message2 = update.message.text
    user_message = replace_standart(user_message2)
    # user_message = standardize_model_name(user_message1, 1)
    user_tag = update.message.from_user.username or update.message.from_user.full_name
    # logger.info("")
    logger.info(f"Получено сообщение от {user_tag}: {user_message}")
    # logger.info("")

    # Если сообщение не является запросом изображения, продолжаем стандартную обработку через GPT
    try:

        query_embedding = create_embedding_for_query(user_message, update)
        most_similar_texts, most_similar_refs, most_similar_related_tables = (
            find_most_similar(query_embedding)
        )

        # Фильтруем и приоритизируем контекст
        prioritized_texts_and_tables, prioritized_images, additional_contexts = (
            filter_and_prioritize_context(
                most_similar_texts, most_similar_refs, most_similar_related_tables
            )
        )

        # Формируем текст контекста из текстов и таблиц
        context_text = "\n\n".join(
            [f"{obj[0]}" for obj in prioritized_texts_and_tables]
            # [f"{obj[0]} ({obj[1]})" for obj in prioritized_texts_and_tables] - закоментил, т.к. после текстового блока было системное имя родительной таблицы
        )

        # Добавляем изображения в контекст (если есть)
        if prioritized_images:
            context_text += "\n\nРисунки и текста:\n" + "\n".join(
                [
                    # f"{img[0]} ({img[1]})" for img in prioritized_images - пока не нужен, img1 - столбик reference в Milvus
                    f"{img[0]}"
                    for img in prioritized_images
                ]  # img[1] теперь берет related_table
            )

        if additional_contexts:
            context_text += "\n\nДополнительный контекст:\n" + "\n".join(
                additional_contexts
            )

        table_contexts = []
        images_to_mention = []
        unique_table_references = set()  # Множество для уникальных ссылок на таблицы

        # Проверяем таблицы и ищем изображения
        for i, ref in enumerate(most_similar_refs):
            if ref.endswith(".xlsx"):  # Если это таблица
                if ref not in unique_table_references:
                    unique_table_references.add(ref)
                    table_content = read_table_from_minio(
                        f"{minio_folder_docs_name}/{ref}"
                    )
                    if table_content:
                        table_name = most_similar_texts[i]
                        table_contexts.append(
                            f"-------\nНачало\n{table_name}:\n{table_content}\nКонец таблицы",
                        )
                        logger.info(f"Использована таблица из MinIO: {ref}")
                    else:
                        logger.warning(f"Пропущена таблица {ref} из-за ошибок чтения.")
            elif re.search(
                r"Рисунок \d+ \(.+\)", most_similar_texts[i]
            ):  # Если это изображение
                images_to_mention.append((most_similar_texts[i], ref))

        if table_contexts:
            context_text += "\n\nТаблицы:\n" + "\n\n".join(table_contexts)

        # Сохраняем контекст в лог-файл
        log_filename = save_context_to_log(user_tag, context_text)
        # Логирование файла для отладки (опционально)
        logger.info(
            f"Контекст для пользователя {user_tag} сохранен в файл: {log_filename}"
        )

        token_count = count_tokens(context_text)
        logger.info(f"Контекст содержит {token_count} токенов")
        # logger.info(f"Используемый контекст: {context_text}")

        # Ищем упоминания рисунков в ответе и создаем ссылки на них
        all_image_mentions = find_image_mentions(context_text)
        """print(f"Проверка 1")
        print(f"{all_image_mentions}")
        print(f"Конец проверки 1")"""
        all_table_mentions = find_table_mentions(context_text)

        images_to_mention = []
        tables_to_mention = []
        for image_text in all_image_mentions:
            """print(f"Проверка 2")
            print(f"{image_text}")
            print(f"Конец проверки 2")"""
            image_ref = find_image_reference_in_milvus(image_text)
            if image_ref:
                images_to_mention.append((image_text, image_ref))

        images_text = "\n".join([img[0] for img in images_to_mention])

        for table_text in all_table_mentions:
            table_ref = find_image_reference_in_milvus(table_text)
            if table_ref:
                tables_to_mention.append((table_text, table_ref))

        # context_text1 = standardize_model_name(context_text, 0)
        logger.info(f"Используемый контекст: {context_text}")
        logger.info("Отправка контекста к GPT")
        # Отправка всего контекста к GPT
        response = openai.chat.completions.create(
            model=MODEL_GPT_INT,
            messages=[
                {
                    "role": "system",
                    "content": (
                        "Я хочу, чтобы ты выступил в роли асистента-помощника для инженеров. "
                        "Твоя основная задача - отвечать на вопросы, анализируя предоставленные данные, без выдумывания информации. Если нужной информации нет, просто скажи, что не можешь ответить на вопрос, так как данных недостаточно."
                        ""
                        "Примечания к контексту:"
                        "Если в контексте будут таблицы, ты должен извлечь из них всю информацию (без вырезания информации), не сжимая ее и отправить эту таблицу в виде списка "
                        'Если в контексте в таблицах узаканы рисунки, ты должен всегда упоминать их все в ответе в формате "Рисункок X" '
                        "Всегда указывай в ответе упомянутые Рисунки (не в конце ответа, а во всем тексте ответа)"
                        # "Если ты упоминаешь рисунки, то упоминай их в формате Рисунок Х."
                        # "Если ты упоминаешь таблицы, то упоминай их в формате ТаблицЕ Х"
                        # "Если ты упоминаешь таблицы, то не склоняй Таблицы\Таблиц\Таблице Х и т.д. Всегда пиши ТаблиЦА Х"
                        "Никогдда не отвечай в виде таблицы, вместо этого отвечай в виде списка"
                        ""
                        ""
                        "Как отвечать:"
                        "Не овечай в духе 'Лучше всего обратиться к производителю или квалифицированному инженеру для получения точной информации'"
                        "Всегда при ответе указывай, на основе каких таблиц(В названии таблицы есть слово 'Таблица ') была основана большая часть твоего ответа, пиши ее/их имя полностью."
                        "не склоняй и не меняй форму названия таблицы, если упоминаешь, то пиши Таблица"
                        "Если нет релевантных изображений/таблиц - Не пиши что 'релевантные изображения/таблицы:отсутствуют' или 'Таблицы, на которых основан ответ:- отсутствуют' если нет таких, то вообще ничего не пиши"
                        "Если тебе запрещено что-то или ты не можешь предоставить, не говори это пользователю"
                        "Не указывай текстовые боки в ответе"
                        "Если в контексте были упомянуты рисунки, то упоминай их так же в своем ответе "
                        ""
                        "Если пользователь запрашивает таблицу (например, 'Таблица Х' или 'Таблица Х полностью' или 'Что находится в Таблице Х', 'Что в Таблице Х')"
                        "ты должен сообщить, что Таблица Х (название) есть в БД, без вывода содержимого таблицы. не говори, что ты не можешь предоставить ее содержимое"
                        "Не отвеча 'Не могу ответить на вопрос, так как данных недостаточно', вместо этого отвечай, что 'Информации не найдено в справочнике'"
                        ""
                        # "Если встречаешь название модели, которое может быть переведено с русского на английский (или наоборот), старайся определить наиболее точное соответствие."
                    ),
                },
                {
                    "role": "system",
                    "content": f"Дополнительные изображения по вашему запросу:\n\n{images_text}\n\n"
                    f"Вот релевантная информация:\n\n{context_text}",
                },
                {"role": "user", "content": user_message},
            ],
            temperature=0.3,
            timeout=30,
        )
        # logger.info(f"response ответа {response}")

        bot_reply = response.choices[0].message.content

        logger.info("Получен ответ от GPT")
        logger.info("Начинается обработка ответа")

        # Найти дополнительные упоминания рисунков, которые есть только в bot_reply
        additional_image_mentions = find_image_mentions(bot_reply)
        additional_table_mentions = find_table_mentions(bot_reply)

        for image_text in additional_image_mentions:
            if image_text not in [mention[0] for mention in images_to_mention]:
                # Если упоминание найдено в bot_reply, но не в контексте, ищем его ссылку
                image_ref = find_image_reference_in_milvus(image_text)
                if image_ref:
                    images_to_mention.append((image_text, image_ref))

        for table_text in additional_table_mentions:
            if table_text not in [mention[0] for mention in tables_to_mention]:
                # Если упоминание найдено в bot_reply, но не в контексте, ищем его ссылку
                table_ref = find_image_reference_in_milvus(table_text)
                if table_ref:
                    tables_to_mention.append((table_text, table_ref))

        bot_reply = response.choices[0].message.content
        # Замена символов < и > на HTML-эквиваленты
        bot_reply = bot_reply.replace("<", "&lt;").replace(">", "&gt;")

        # print("Список images_to_mention")
        # print(images_to_mention)
        # print("Конец списка images_to_mention")
        bot_reply = normalize_mentions(bot_reply)
        formatted_reply = format_image_links(bot_reply, images_to_mention)
        logger.info(f"Отправка ответа пользователю {user_tag}: {formatted_reply}")
        await send_large_message(update, formatted_reply)
        await send_table_to_chat(update, tables_to_mention, formatted_reply)
        await request_feedback(update, context)

        images_to_send = []
        for image_text, ref in images_to_mention:
            if image_text.split(" ")[0] in bot_reply:
                images_to_send.append(ref)

        save_user_question_to_sheet(
            user_message, bot_reply, user_tag, log_filename, "Режим Справочника"
        )

        await asyncio.sleep(1)
    except Exception as e:
        error_message = str(e)
        logger.error(f"Произошла ошибка: {e}")
        # 1️⃣ Проверяем, произошла ли ошибка с несовпадающими размерностями (ошибка в создании вектора вопроса)
        if "shapes" in error_message and "not aligned" in error_message:
            await update.message.reply_text(
                "⚠️ Ошибка при векторизации вопроса!\n"
                "Пожалуйста, сообщите администратору."
            )

        # 2️⃣ Проверяем, является ли ошибка 403 (регион не поддерживается) (VPN ЛЁГ)
        elif "unsupported_country_region_territory" in error_message:
            await update.message.reply_text(
                "⚠️ Ошибка с доступом к OpenAI API для вашего региона.\n"
                "Пожалуйста, сообщите администратору."
            )

        # 3️⃣ Если ошибка другая — обычный вывод
        else:
            await update.message.reply_text(
                f"❌ Произошла ошибка при получении ответа:\n{error_message}"
            )


# нормализует слово standard
def replace_standart(text):
    """
    Заменяет все варианты слова 'Standart' (Standart, STANDART, standart) на 'Standard'.

    Аргументы:
        text (str): Входной текст.

    Возвращает:
        str: Текст с заменёнными словами.
    """
    print("запустился метод replace_standart")
    return re.sub(r"\b[Ss][Tt][Aa][Nn][Dd][Aa][Rr][DdTt]\b", "Standard", text)


# добавляет пропуски в названиях, убирая дефис и разъединяя буквы и цифры
def standardize_model_name(model_name, param):
    # Добавляем пробел перед цифрами, если его нет
    # model_name = re.sub(r"([A-Za-z]+)(\d+)", r"\1 \2", model_name)
    # Заменяем только те тире, которые находятся между буквами или цифрами
    if param == 1:
        model_name = re.sub(r"([A-Za-z]+)(\d+)", r"\1 \2", model_name)
    model_name = re.sub(r"(?<=[A-Za-z0-9])-(?=[A-Za-z0-9])", " ", model_name)
    return model_name


# Метод для преобразования склонений упомянутых таблиц и рисунков
def normalize_mentions(gpt_response):
    """
    Исправляет склонения слов 'Рисунок' и 'Таблица' на базовые формы перед числами.
    """
    # print(f"gpt ответ до исправлений: {gpt_response}")
    # Шаблон для склонений "Рисунок" перед числами
    pattern_risunok = r"[Рр]исунк[аеуов]{1}(?=\s*\d+)"
    # Шаблон для склонений "Таблица" перед числами
    pattern_tablitsa = r"[Тт]аблиц[аеуовы]{1}(?=\s*\d+)"

    pattern_risunok2 = r"[Рр]исунок{1}(?=\s*\d+)"
    pattern_tablitsa2 = r"[Тт]аблица{1}(?=\s*\d+)"

    # Заменяем склонения "Рисунок" на базовую форму
    gpt_response = re.sub(pattern_risunok, "Рисунок", gpt_response)
    gpt_response = re.sub(pattern_risunok2, "Рисунок", gpt_response)
    # Заменяем склонения "Таблица" на базовую форму
    gpt_response = re.sub(pattern_tablitsa, "Таблица", gpt_response)
    gpt_response = re.sub(pattern_tablitsa2, "Таблица", gpt_response)

    # Логируем результат
    # print(f"gpt ответ после исправлений: {gpt_response}")

    return gpt_response


# Метод для обработки сообщений в режиме мануалов
async def handle_message_manuals(update: Update, context):
    # print("Заработал режим handle_message_manuals")
    if context.user_data.get("last_selected_mode") != "manuals_engrs":
        logger.error("handle_message_manuals вызван вне режима мануалов.")
        return

    user_id = update.message.from_user.id

    # Проверяем, ждет ли бот оценку
    if context.user_data.get("awaiting_feedback", False):
        await update.message.reply_text(
            "⚠️ Сначала оцените предыдущий ответ, прежде чем задать новый вопрос!"
        )
        return  # Блокируем новый вопрос

    user_message = update.message.text
    user_tag = update.message.from_user.username or update.message.from_user.full_name
    # print("Точка1")
    try:

        query_embedding = create_embedding_for_query(user_message, update)
        # print(f"query_embedding - {query_embedding}")

        # Получаем релевантные тексты и коллекции
        """related_collections = find_most_similar_with_collections(
            context, query_embedding
        )"""
        # print("Точка 1")
        related_collections, related_texts, related_score = search_in_milvus(
            query_embedding
        )

        # print(f"Список релевантных мануалов:")

        responce = f"Найдены документы, схожие по отрывкам\n"
        count_finds = 1
        for col, txt, score in zip(related_collections, related_texts, related_score):
            responce += (
                f"{count_finds}) Релевантный документ - {col}\n"  # Название коллекции
            )
            responce += f"Релевантный текст - {txt}\n"  # Текст
            responce += f"Векторное совпадение - {score}%\n\n"  # Текст
            count_finds += 1
        await send_large_message_for_manuals(update, responce)
        await request_feedback(update, context)

        # Сохраняем контекст в лог-файл
        log_filename = save_context_to_log(user_tag, responce)
        # Логирование файла для отладки (опционально)
        logger.info(
            f"Контекст для пользователя {user_tag} сохранен в файл: {log_filename}"
        )

        # Логируется в табличку
        save_user_question_to_sheet(
            user_message, responce, user_tag, log_filename, "Рижим Мануалов"
        )
        # await update.message.reply_text(responce)

        """for col in related_collections:
            responce += col + "\n"
            # await update.message.reply_text(col)

        for text in related_texts:
            responce += "Релевантный текст - " + text + "\n"
            # await update.message.reply_text(col)
        await update.message.reply_text(responce)

        additional_table_mentions = find_table_mentions(responce)"""

    except Exception as e:
        logger.error(f"Ошибка обработки сообщения в режиме мануалов: {e}")
        await update.message.reply_text("Произошла ошибка при обработке запроса.")


# Метод поиска упомянутых изображений по формату "Рисунок Х"
def search_by_figure_id(figure_id):
    collection = Collection(name=milvus_collection_name)
    try:
        result = collection.query(
            expr=f'figure_id == "{figure_id.strip()}"',  # Удаляем лишние пробелы
            output_fields=["text", "reference"],
        )
        if result:
            return result[0]["text"]
    except Exception as e:
        logger.error(f"Ошибка при поиске в Milvus для '{figure_id}': {e}")
    return None


# Метод добавляет ссылки на упомянутые изображения в ответе GPT
def format_image_links(bot_reply, images_to_mention):
    """Форматирует текст ответа, добавляя кликабельные ссылки на изображения."""
    for image_text, ref in images_to_mention:
        # Создаем URL для изображения
        image_url = (
            f"{MINIO_ENDPOINT}/{MINIO_BUCKET_NAME}/{minio_folder_docs_name}/{ref}"
        )
        # print(f"{image_url}, {ref}")
        # logger.info(f"найденные все картинки - {image_text} {ref}")
        # Формируем кликабельную ссылку в формате HTML
        link_text = f'<a href="{image_url}" target="_blank">{image_text}</a>'
        # print("Проверка link_text")
        # print(link_text)
        # Заменяем все упоминания "Рисунок X" на кликабельную ссылку
        bot_reply = re.sub(
            rf"\b{re.escape(image_text)}\b",  # \b обеспечивает точное совпадение слова
            link_text,
            bot_reply,
        )

    return bot_reply


# Метод, находящий в MiniO таблички по упоминанию "Таблица Х"
async def send_table_to_chat(update, tables_to_mention, formatted_reply):
    """
    Находит таблицы в MinIO по упоминанию, проверяет их присутствие в ответе GPT,
    исключает повторную отправку и отправляет их в чат Telegram.
    """
    sent_tables = set()  # Хранилище для уже отправленных таблиц

    for table_text, ref in tables_to_mention:
        # Проверяем, упоминается ли таблица в ответе GPT
        # Используем регулярное выражение для точного совпадения таблицы
        pattern = rf"\b{re.escape(table_text)}\b"  # \b обозначает границы слова

        if not re.search(
            pattern, formatted_reply
        ):  # Если таблица не упоминается, пропускаем
            continue

        # Проверяем, отправлялась ли таблица ранее
        if ref in sent_tables:
            # logger.info(f"Таблица {table_text} уже была отправлена ранее. Пропускаем.")
            continue

        logger.info(f"Обработка таблицы: {table_text} с системным именем {ref}")
        try:
            # Проверяем существование таблицы в MinIO
            table_key = f"{minio_folder_docs_name}/{ref}"
            response = s3_client.get_object(Bucket=MINIO_BUCKET_NAME, Key=table_key)
            file_data = response["Body"].read()

            # Отправляем таблицу пользователю как документ
            await update.message.reply_document(
                document=BytesIO(file_data),
                filename=f"{table_text}.xlsx",
                # caption=f"Таблица {table_text} из вашего запроса.",
            )
            # logger.info(f"Таблица {table_text} успешно отправлена.")

            # Добавляем таблицу в список отправленных
            sent_tables.add(ref)
        except Exception as e:
            logger.error(f"Не удалось отправить таблицу {table_text}: {e}")
            await update.message.reply_text(
                f"Ошибка при отправке таблицы {table_text}."
            )


# Метод, разделяющий сообщения от ТГ Бота по 4000 символов с лог заглючением по абзацам
async def send_large_message(update, text, max_length=4000):
    # Разбиваем текст по абзацам
    paragraphs = text.split("\n\n")
    current_message = ""

    for paragraph in paragraphs:
        # Проверяем, если текущий абзац слишком длинный, чтобы отправить его как есть
        if len(paragraph) > max_length:
            # Если абзац превышает max_length, разбиваем его на подчасти
            sub_paragraphs = [
                paragraph[i : i + max_length]
                for i in range(0, len(paragraph), max_length)
            ]
            for sub_paragraph in sub_paragraphs:
                # await update.message.reply_text(sub_paragraph)
                await update.message.reply_text(sub_paragraph, parse_mode="HTML")
            continue  # Переходим к следующему абзацу после отправки разбиения

        # Проверяем, можно ли добавить текущий абзац в сообщение
        if len(current_message) + len(paragraph) + 2 <= max_length:
            # Добавляем абзац в текущее сообщение
            if current_message:
                current_message += "\n\n" + paragraph
            else:
                current_message = paragraph
        else:
            # Если текущее сообщение заполнено, отправляем его и начинаем новое
            # await update.message.reply_text(current_message)
            await update.message.reply_text(current_message, parse_mode="HTML")
            current_message = paragraph  # Начинаем новое сообщение с текущего абзаца

    # Отправляем оставшуюся часть сообщения, если что-то осталось
    if current_message:
        # await update.message.reply_text(current_message)
        await update.message.reply_text(current_message, parse_mode="HTML")


async def send_large_message_for_manuals(update, text, max_length=4000):
    # Разбиваем текст по абзацам
    paragraphs = text.split("\n\n")
    current_message = ""

    for paragraph in paragraphs:
        # Проверяем, если текущий абзац слишком длинный, чтобы отправить его как есть
        if len(paragraph) > max_length:
            # Если абзац превышает max_length, разбиваем его на подчасти
            sub_paragraphs = [
                paragraph[i : i + max_length]
                for i in range(0, len(paragraph), max_length)
            ]
            for sub_paragraph in sub_paragraphs:
                await update.message.reply_text(sub_paragraph)
                # await update.message.reply_text(sub_paragraph, parse_mode="HTML")
            continue  # Переходим к следующему абзацу после отправки разбиения

        # Проверяем, можно ли добавить текущий абзац в сообщение
        if len(current_message) + len(paragraph) + 2 <= max_length:
            # Добавляем абзац в текущее сообщение
            if current_message:
                current_message += "\n\n" + paragraph
            else:
                current_message = paragraph
        else:
            # Если текущее сообщение заполнено, отправляем его и начинаем новое
            await update.message.reply_text(current_message)
            # await update.message.reply_text(current_message, parse_mode="HTML")
            current_message = paragraph  # Начинаем новое сообщение с текущего абзаца

    # Отправляем оставшуюся часть сообщения, если что-то осталось
    if current_message:
        await update.message.reply_text(current_message)
        # await update.message.reply_text(current_message, parse_mode="HTML")


# Метод дополнительного поиск упомянутых изображений в ответе GPT по Рисунок Х
def find_image_mentions(text):
    pattern = r"Рисунок \d+"

    return re.findall(pattern, text)


# Метод дополнительного поиск упомянутых таблиц в ответе GPT по Таблица ...
def find_table_mentions(text):
    pattern = r"Таблица \d+"  # Ищет фразы, начинающиеся с "Таблица"
    return re.findall(pattern, text)


def find_image_reference_in_milvus(figure_id):
    collection = Collection(name=milvus_collection_name)
    try:
        result = collection.query(
            expr=f'figure_id == "{figure_id}"', output_fields=["reference"]
        )
        # print("Проверка 1")
        # print(f"figure_id - {figure_id}")
        # print(f"Найденный результат - {result[0]["reference"]}")
        if result:
            return result[0]["reference"]
    except Exception as e:
        logger.error(f"Ошибка при поиске в Milvus для '{figure_id}': {e}")
    return None


# Метод доработка лог файла с контекстом пользователя
def sanitize_filename(filename):
    """Функция для удаления или замены недопустимых символов в названии файла."""
    return "".join(c if c.isalnum() or c in (" ", "_", "-") else "_" for c in filename)


# Метод создает лог файл с контекстом, отправляемым на основе вопроса юзера
def get_unique_log_filename(user_tag):
    # Создаем уникальное имя файла на основе временной метки и никнейма пользователя
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    sanitized_tag = sanitize_filename(
        user_tag
    )  # Очищаем никнейм от недопустимых символов
    return f"context_log_{sanitized_tag}_{timestamp}.txt"


# Метод создания лог файл
def save_context_to_log(user_tag, context_text):
    # Генерируем уникальное имя для лог-файла
    unique_log_filename = get_unique_log_filename(user_tag)
    log_key = f"{MINIO_FOLDER_LOGS_NAME}/{unique_log_filename}"  # Лог будет храниться в бакете под ключом logs/имя_файла

    try:
        # Сохраняем лог в MinIO
        s3_client.put_object(
            Bucket=MINIO_BUCKET_NAME,  # Имя бакета из переменной окружения
            Key=log_key,  # Путь (ключ) к файлу в бакете
            Body=context_text.encode("utf-8"),  # Содержимое файла
        )
        # logger.info(f"Файл {unique_log_filename} успешно сохранён в бакете MinIO")
    except Exception as e:
        logger.error(f"Ошибка сохранения файла {unique_log_filename} в MinIO: {e}")
        raise

    return log_key  # Возвращаем ключ файла в бакете вместо локального пути


# Метод для обработки оценок
async def handle_feedback(update: Update, context):
    quality_score = update.message.text  # Получение оценки пользователя
    next_row = len(sheet.get_all_values())  # Нахождение строки для записи оценки
    sheet.update(f"D{next_row}", [[quality_score]])  # Запись оценки в 4-й столбик
    await update.message.reply_text(reply_markup=ReplyKeyboardRemove())


# Метод отчищает сообщения, полученные в момент отключения
def clear_message_bot():

    # Установка offset, чтобы удалить все накопленные сообщения
    response = requests.get(URL)
    if response.status_code == 200:
        updates = response.json()

        # Проверяем наличие ключа 'result' и его содержимое
        if "result" in updates and updates["result"]:
            for update in updates["result"]:
                # Извлекаем данные
                username = (
                    update.get("message", {})
                    .get("from", {})
                    .get("username", "Неизвестный пользователь")
                )
                text = update.get("message", {}).get("text", "Без текста")

                # Логируем только нужную информацию
                logger.info(f"Пользователь {username} отправил сообщение: {text}")

            # Очищаем очередь сообщений
            last_update_id = updates["result"][-1]["update_id"]
            clear_url = f"{URL}?offset={last_update_id + 1}"
            requests.get(clear_url)
            logger.info("Очередь сообщений очищена.")
        else:
            logger.info("Нет новых сообщений.")
    else:
        logger.info(f"Ошибка API Telegram: {response.status_code}, {response.text}")


"""# Обработка выбора базы данных через callback кнопки в ТГ Боте
async def select_db(update: Update, context: ContextTypes.DEFAULT_TYPE):
    print("Запустился метод select_db")
    query = update.callback_query
    context.user_data["minio_folder_docs_name"] = MINIO_FOLDER_DOCS_NAME_SPRAVOCHNIK
    await query.answer()  # Подтверждаем получение запроса

    # global minio_folder_docs_name  # Объявляем переменную глобальнойы
    global milvus_collection_name  # Объявляем переменную глобальной

    selected_db = query.data  # Получаем callback_data из нажатой кнопки
    context.user_data["selected_db"] = selected_db  # Сохраняем выбор пользователя
    # connect_to_milvus(selected_db)  # Подключаемся к выбранной базе данных
    connections.disconnect(alias="default")  # Отключаемся от нынешней бд в Milvus

    # Отправляем сообщение пользователю
    if selected_db == "engs_bot":
        await query.edit_message_text(f"Вы выбрали режим справочника: {selected_db}")
        minio_folder_docs_name = MINIO_FOLDER_DOCS_NAME_SPRAVOCHNIK  # Изменение папки для поиска таблиц\рисунков
        # logger.info(minio_folder_docs_name)
        milvus_collection_name = MILVUS_COLLECTION  # Изменение коллекции milvus
        # logger.info(milvus_collection_name)
        context.user_data["minio_folder_docs_name"] = MINIO_FOLDER_DOCS_NAME_SPRAVOCHNIK
        context.user_data["milvus_collection_name"] = MILVUS_COLLECTION
        context.user_data["handle_message_method"] = handle_message
        # Подключаемся к Milvus с справочником
        connections.connect(
            alias="default",
            host=MILVUS_HOST,
            port=MILVUS_PORT,
            db_name=MILVUS_DB_NAME_FIRST,
            user=MILVUS_USER,
            password=MILVUS_PASSWORD,
        )
    else:
        await query.edit_message_text(
            f"Вы выбрали режим поиска мануалов: {selected_db}"
        )
        minio_folder_docs_name = (
            MINIO_FOLDER_DOCS_NAME_MANUAL  # Изменение папки для поиска таблиц\рисунков
        )
        # logger.info(minio_folder_docs_name)
        milvus_collection_name = MILVUS_COLLECTION  # Изменение коллекции milvus
        # logger.info(milvus_collection_name)
        context.user_data["minio_folder_docs_name"] = MINIO_FOLDER_DOCS_NAME_MANUAL
        context.user_data["milvus_collection_name"] = MILVUS_COLLECTION
        context.user_data["handle_message_method"] = handle_message_manuals
        connections.connect(
            alias="default",
            host=MILVUS_HOST,
            port=MILVUS_PORT,
            db_name=MILVUS_DB_NAME_FIRST,
            user=MILVUS_USER,
            password=MILVUS_PASSWORD,
        )
    await context
    # Загружаем коллекции при втором режиме работы"""


# Метод для получения описания (description) коллекции Milvus
def get_collection_description(collection_name):
    # logger.error(f"Вызвался метод get_collection_description!!!")
    try:
        collection = Collection(name=collection_name)
        return collection.description  # Возвращаем описание коллекции
    except Exception as e:
        logger.error(
            f"Не удалось получить описание для коллекции '{collection_name}': {e}"
        )
        return None


async def set_bot_commands(application):
    """
    Устанавливает меню команд для Telegram-бота.
    """
    commands = [
        BotCommand("start", "Запустить бота"),
        # BotCommand("help", "Получить помощь"),
        BotCommand("metod", "Выбрать режим работы бота"),
        BotCommand("info", "Информация о боте"),
    ]
    await application.bot.set_my_commands(commands)


async def request_feedback(update, context):
    """
    Метод предлагает пользователю выбрать оценку ответа с помощью Inline-кнопок.
    """
    keyboard = [
        [InlineKeyboardButton("Хорошо 🟢", callback_data="feedback_good")],
        [
            InlineKeyboardButton(
                "Удовлетворительно 🟡", callback_data="feedback_neutral"
            )
        ],
        [InlineKeyboardButton("Плохо 🔴", callback_data="feedback_bad")],
    ]

    reply_markup = InlineKeyboardMarkup(keyboard)

    await update.message.reply_text(
        "📢 Пожалуйста, оцените ответ перед тем, как задать новый вопрос:",
        reply_markup=reply_markup,
    )
    context.user_data["awaiting_feedback"] = True  # Блокируем следующий вопрос


async def handle_all_callbacks(update: Update, context):
    """Обрабатывает все CallbackQuery и перенаправляет в нужный обработчик."""
    query = update.callback_query
    await query.answer()  # Подтверждаем получение нажатия

    # Определяем, какая кнопка была нажата
    if query.data in ["engs_bot", "manuals_engrs"]:
        await handle_callback_metod(update, context)  # Вызов выбора режима
    elif query.data.startswith("feedback_"):
        await handle_feedback_callback(update, context)  # Вызов обработки оценки
    else:
        logger.warning(f"Неизвестный callback_data: {query.data}")


async def handle_feedback_callback(update: Update, context):
    print("вызван метод handle_feedback_callback")
    """Обрабатывает нажатие на кнопки оценки ответа."""

    query = update.callback_query
    await query.answer()  # Подтверждаем нажатие

    # Словарь с вариантами оценок
    feedback_map = {
        "feedback_good": "Хорошо 🟢",
        "feedback_neutral": "Удовлетворительно 🟡",
        "feedback_bad": "Плохо 🔴",
    }
    feedback_text = feedback_map.get(query.data, "Неизвестная оценка")

    # Разрешаем пользователю задавать новый вопрос
    context.user_data["awaiting_feedback"] = False  # Снимаем блокировку

    user_tag = (
        query.from_user.username or query.from_user.full_name
    )  # Получаем user_tag

    # Получаем все данные из Google Таблицы
    all_data = sheet.get_all_values()

    # Поиск строки с user_tag (колонка E) и пустой оценкой (колонка D)
    row_index = None
    for i in range(len(all_data) - 1, 0, -1):  # Проходим с конца к началу
        if len(all_data[i]) >= 5:  # Проверяем, что строка имеет хотя бы 5 колонок
            if (
                all_data[i][4] == user_tag and all_data[i][3] == ""
            ):  # user_tag в E, оценка в D пуста
                row_index = i + 1  # Google Sheets использует индексацию с 1
                break

    if row_index:
        # Записываем оценку в колонку D
        sheet.update(f"D{row_index}", [[feedback_text]])
        logger.info(
            f"✅ Оценка '{feedback_text}' записана для {user_tag} в строку {row_index}"
        )
        await query.edit_message_text(f"Вы выбрали оценку: {feedback_text}")
    else:
        logger.warning(f"⚠ Не найден вопрос без оценки для {user_tag}")
        await query.edit_message_text(f"⚠ Не найден ваш вопрос для оценки.")


# Метод обработки ошибки асинхронной менюшки /comands
def run_async_task(task):
    try:
        loop = asyncio.get_running_loop()
    except RuntimeError:  # Если цикла нет, создаем его
        loop = asyncio.new_event_loop()
        asyncio.set_event_loop(loop)
    return loop.run_until_complete(task)


def save_user_mode_to_sheet(user_id, mode):
    """Сохраняет выбранный режим работы пользователя в 3-й лист Google Sheets."""
    try:
        worksheet = client.open_by_key(SPREADSHEET_ID).worksheet(
            "Ласт выбранный метод пользователей"
        )
        all_data = worksheet.get_all_values()  # Получаем все строки

        if not all_data:  # Если лист вообще пустой
            worksheet.append_row(["Telegram ID", "Метод работы"])  # Добавляем заголовки

        user_ids = [
            row[0] for row in all_data[1:] if row
        ]  # Получаем ID, пропуская заголовки

        if str(user_id) in user_ids:
            row_index = (
                user_ids.index(str(user_id)) + 2
            )  # Индекс в Google Sheets (начинается с 1)
            worksheet.update(f"B{row_index}", [[mode]])  # Обновляем режим работы
        else:
            worksheet.append_row([str(user_id), mode])  # Добавляем новую запись

        logger.info(f"Режим работы '{mode}' сохранен для пользователя {user_id}")

    except Exception as e:
        logger.error(f"Ошибка при сохранении режима работы в Google Sheets: {e}")


def load_user_mode_from_sheet(user_id):
    """Загружает последний выбранный режим работы пользователя из 3-го листа Google Sheets."""
    try:
        worksheet = client.open_by_key(SPREADSHEET_ID).worksheet(
            "Ласт выбранный метод пользователей"
        )
        all_data = worksheet.get_all_values()

        if (
            not all_data or len(all_data) < 2
        ):  # Проверка, есть ли данные (пропускаем заголовки)
            return None

        for row in all_data[1:]:  # Пропускаем заголовок
            if len(row) >= 2 and row[0] == str(
                user_id
            ):  # Проверяем, есть ли вторая колонка
                return row[1]  # Возвращаем метод работы

        return None  # Если ID не найден
    except Exception as e:
        logger.error(f"Ошибка при загрузке режима работы из Google Sheets: {e}")
        return None


# метод для обработки нажатой кнопки при выбор режима работы Бота
async def handle_callback_metod(update: Update, context: ContextTypes.DEFAULT_TYPE):
    # logger.info("Запустился метод handle_callback_metod")
    query = update.callback_query
    query.answer()  # Подтверждаем получение нажатия

    user_id = query.from_user.id
    selected_method = query.data  # Получаем, что выбрал пользователь

    # Проверяем, какую кнопку нажали
    if selected_method == "engs_bot":
        await query.edit_message_text(
            "Вы выбрали режим : Поиск по справочнику📔 \n\nМожете задать вопрос"
        )
        context.user_data["handle_message_method"] = handle_message

    elif selected_method == "manuals_engrs":
        await query.edit_message_text(
            "Вы выбрали режим: Поиск мануалов📚 \n\nМожете задать вопрос"
        )
        context.user_data["handle_message_method"] = handle_message_manuals

    # Сохраняем метод работы в память бота
    context.user_data["selected_method"] = selected_method

    # Сохраняем в Google Sheets (3-й лист)
    save_user_mode_to_sheet(user_id, selected_method)


async def handle_message_async(update: Update, context):
    asyncio.create_task(handle_message(update, context))  # Запускаем как задачу


# Основная функция для запуска бота
def main():
    application = ApplicationBuilder().token(TELEGRAM_BOT_TOKEN).build()
    # Установка команд для меню
    run_async_task(set_bot_commands(application))

    application.add_handler(CommandHandler("start", start))  # Обработка команды /start
    application.add_handler(CommandHandler("info", info))  # Обработка команды /info
    application.add_handler(CommandHandler("metod", metod))  # Обработка команды /metod
    application.add_handler(
        CallbackQueryHandler(handle_all_callbacks)
    )  # оббработка нажатия кнопок по выбору режма работы Бота

    application.add_handler(MessageHandler(filters.TEXT, handle_message))

    # application.add_handler(MessageHandler(filters.TEXT, handle_message_manuals))

    # Метот обработки после нажатия кнопки оценки ответа
    application.add_handler(MessageHandler(filters.TEXT, handle_feedback))

    logger.info("Бот запущен.")
    clear_message_bot()
    application.run_polling()
    """application.run_webhook(
        listen="localhost", port=80, webhook_url="https://exapmle.com:80"
    )  # Тест через JMeter. Включаешь это и выключаешь "application.run_polling()"
    """


if __name__ == "__main__":
    main()
